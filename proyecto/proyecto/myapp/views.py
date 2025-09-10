from django.shortcuts import render
from django.http import HttpResponse 
from .models import Zona, Espacio, EstadoEspacio, Responsable, TipoActividad, EspacioRecurso, Reserva, Usuario, Recurso, TipoReserva
from django.db import connection
from django.db.models import Count, Q, Value, CharField, Case, When, IntegerField
from django.db.models.functions import Concat
from django.db.models import Prefetch, Exists, OuterRef
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from datetime import date, timedelta, time, datetime
import pytz
from openpyxl import Workbook
import openpyxl
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import authenticate
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth.decorators import login_required
import jwt
from django.conf import settings
from django.shortcuts import redirect
from django.utils import timezone
from django.contrib.auth import login as auth_login
import pandas as pd



def token_required(view_func):
    def wrapper(request, *args, **kwargs):
        token = request.session.get("access_token")
        if not token:
            return redirect("/login")

        try:
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=["HS256"])
            exp = timezone.datetime.fromtimestamp(payload["exp"], tz=timezone.utc)
            if timezone.now() > exp:
                return redirect("/login")
        except jwt.ExpiredSignatureError:
            return redirect("/login")
        except jwt.InvalidTokenError:
            return redirect("/login")

        return view_func(request, *args, **kwargs)
    return wrapper


def disponibilidad_espacios(request):
    tz = pytz.timezone('America/Santiago')
    now = timezone.now().astimezone(tz)
    current_date = now.date()
    current_time = now.time()

    zonas = Zona.objects.prefetch_related(
        Prefetch('espacio_set', 
                 queryset=Espacio.objects.annotate(
                     tiene_reserva_activa=Exists(
                         Reserva.objects.filter(
                             idespacio=OuterRef('pk'),
                             fechareserva=current_date,
                             horainicio__lte=current_time,
                             horafin__gte=current_time
                         )
                     )
                 ).select_related('idestadoespacio')
        )
    ).all()

    total_disponibles = 0
    zonas_data = []

    for zona in zonas:
        espacios_data = []
        disponibles = 0
        
        for espacio in zona.espacio_set.all():
            if espacio.idestadoespacio.descripcionestadoespacio == "En mantención":
                estado = "mantencion"
            else:
                if hasattr(espacio, 'tiene_reserva_activa') and espacio.tiene_reserva_activa:
                    estado = "ocupado"
                else:
                    estado = "disponible"
                    disponibles += 1
            
            espacios_data.append({
                'espacio': espacio,
                'estado': estado
            })

        total_disponibles += disponibles

        zonas_data.append({
            'zona': zona,
            'espacios_data': espacios_data,
            'disponibles': disponibles
        })

    return render(request, 'espacios.html', {
        'zonas_data': zonas_data,
        'total_disponibles': total_disponibles
    })


def disponibilidad(request):
    zonas = Zona.objects.prefetch_related('espacio_set__idestadoespacio').all()
    return render(request, 'espacios.html', {'zonas': zonas})

@login_required(login_url='/login')
def personal(request):
    busqueda = request.GET.get('busqueda', '')
    tipo_actividad_id = request.GET.get('tipo_actividad', '')
    
    responsables = Responsable.objects.select_related('idtipoactividad').all().order_by('apellidoresponsable')
    
    if busqueda:
        responsables = responsables.filter(
            Q(nombreresponsable__icontains=busqueda) | 
            Q(apellidoresponsable__icontains=busqueda)
        )
    
    if tipo_actividad_id:
        responsables = responsables.filter(idtipoactividad__idtipoactividad=tipo_actividad_id)
        tipo_actividad_nombre = TipoActividad.objects.get(
            idtipoactividad=tipo_actividad_id
        ).nombretipoactividad
    else:
        tipo_actividad_nombre = None
    
    tipos_actividad = TipoActividad.objects.filter(
        idtipoactividad__in=Responsable.objects.values('idtipoactividad').distinct()
    ).order_by('nombretipoactividad')
    
    return render(request, 'personal.html', {
        'responsables': responsables,
        'tipos_actividad': tipos_actividad,
        'busqueda_actual': busqueda,
        'tipo_actividad_actual': tipo_actividad_id,
        'tipo_actividad_nombre': tipo_actividad_nombre
    })

def cambiar_estado_recurso(request, recurso_id, espacio_id, nuevo_estado):
    try:
        espacio_recurso = get_object_or_404(
            EspacioRecurso,
            idrecurso=recurso_id,
            idespacio=espacio_id
        )
        
        if espacio_recurso.cambiar_estado(nuevo_estado):
            messages.success(request, "Estado del recurso actualizado")
        else:
            messages.error(request, "No se pudo cambiar el estado")
            
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
    
    return redirect('recursos.html')

@login_required(login_url='/login')
def agenda(request):
    ahora_local = timezone.localtime(timezone.now())
    fecha_hoy = ahora_local.date()
    hora_actual = ahora_local.time()
    espacios = Espacio.objects.all()
    tipos_reserva = TipoReserva.objects.all()
    
    mostrar_actuales = 'mostrar_actuales' in request.GET
    
    filtro_responsable = request.GET.get('responsable', '')
    filtro_usuario = request.GET.get('usuario', '')
    filtro_espacio = request.GET.get('espacio', '')
    filtro_tipo_reserva = request.GET.get('tipo_reserva', '')
    
    if mostrar_actuales:
        fecha_seleccionada = fecha_hoy
    else:
        try:
            fecha_seleccionada = date.fromisoformat(request.GET.get('fecha', fecha_hoy.isoformat()))
        except ValueError:
            fecha_seleccionada = fecha_hoy
    
    reservas = Reserva.objects.select_related(
        'rutresponsable', 'rutusuario', 'idespacio', 'idtiporeserva'
    ).filter(
        fechareserva=fecha_seleccionada
    ).order_by('horainicio')
    
    if filtro_responsable:
        reservas = reservas.filter(
            Q(rutresponsable__nombreresponsable__icontains=filtro_responsable) |
            Q(rutresponsable__apellidoresponsable__icontains=filtro_responsable)
        )
    
    if filtro_usuario:
        reservas = reservas.filter(
            Q(rutusuario__nombreusuario__icontains=filtro_usuario) |
            Q(rutusuario__apellidousuario__icontains=filtro_usuario)
        )
    
    if filtro_espacio:
        reservas = reservas.filter(idespacio__idespacio=filtro_espacio)
    
    if filtro_tipo_reserva:
        reservas = reservas.filter(idtiporeserva__idtiporeserva=filtro_tipo_reserva)
    
    if mostrar_actuales:
        reservas = reservas.filter(horainicio__gte=hora_actual)
    
    return render(request, 'agenda.html', {
        'reservas': reservas,
        'fecha_seleccionada': fecha_seleccionada,
        'hoy': fecha_hoy,
        'hora_actual': hora_actual,
        'mostrar_actuales': mostrar_actuales,
        'espacios': espacios,
        'tipos_reserva': tipos_reserva
    })

@login_required(login_url='/login')
def espacio_detail(request, espacio_id):
    ahora_local = timezone.localtime(timezone.now())
    fecha_hoy = ahora_local.date()
    hora_actual = ahora_local.time()

    espacio = get_object_or_404(Espacio, idespacio=espacio_id)

    tipo_actividad = espacio.tipoactividadespacio
    recursos = Recurso.objects.filter(
        espaciorecurso__idespacio=espacio.idespacio
    ).values_list('nombrerecurso', flat=True)
    estado = espacio.idestadoespacio.descripcionestadoespacio

    tipos_reserva = TipoReserva.objects.all()

    mostrar_actuales = 'mostrar_actuales' in request.GET

    filtro_responsable = request.GET.get('responsable', '')
    filtro_usuario = request.GET.get('usuario', '')
    filtro_tipo_reserva = request.GET.get('tipo_reserva', '')

    if mostrar_actuales:
        fecha_seleccionada = fecha_hoy
    else:
        try:
            fecha_seleccionada = date.fromisoformat(request.GET.get('fecha', fecha_hoy.isoformat()))
        except ValueError:
            fecha_seleccionada = fecha_hoy

    reservas = Reserva.objects.select_related(
        'rutresponsable', 'rutusuario', 'idespacio', 'idtiporeserva'
    ).filter(
        fechareserva=fecha_seleccionada,
        idespacio=espacio
    ).order_by('horainicio')

    if filtro_responsable:
        reservas = reservas.filter(
            Q(rutresponsable__nombreresponsable__icontains=filtro_responsable) |
            Q(rutresponsable__apellidoresponsable__icontains=filtro_responsable)
        )

    if filtro_usuario:
        reservas = reservas.filter(
            Q(rutusuario__nombreusuario__icontains=filtro_usuario) |
            Q(rutusuario__apellidousuario__icontains=filtro_usuario)
        )

    if filtro_tipo_reserva:
        reservas = reservas.filter(idtiporeserva__idtiporeserva=filtro_tipo_reserva)

    if mostrar_actuales:
        reservas = reservas.filter(horainicio__gte=hora_actual)

    return render(request, 'espacio.html', {
        'reservas': reservas,
        'fecha_seleccionada': fecha_seleccionada,
        'hoy': fecha_hoy,
        'hora_actual': hora_actual,
        'mostrar_actuales': mostrar_actuales,
        'espacio': espacio,
        'espacio_id': espacio.idespacio,
        'n_espacio': espacio.numeroespacio,
        'tipo_actividad': tipo_actividad,
        'recursos': recursos,
        'estado': estado,
        'tipos_reserva': tipos_reserva
    })

@login_required(login_url='/login')
def recursos(request):
    recursos_data = EspacioRecurso.objects.values(
        'idrecurso',
        'idrecurso__nombrerecurso'
    ).annotate(
        total=Count('idrecurso'),
        disponible=Count(
            Case(
                When(idestadorecurso__descripcion='Disponible', then=1),
                output_field=IntegerField()
            )
        ),
        reparacion=Count(
            Case(
                When(idestadorecurso__descripcion='En reparación', then=1),
                output_field=IntegerField()
            )
        ),
        fuera_servicio=Count(
            Case(
                When(idestadorecurso__descripcion='Fuera de servicio', then=1),
                output_field=IntegerField()
            )
        )
    ).order_by('idrecurso__nombrerecurso')

    recursos_dict = {
        item['idrecurso']: {
            'nombre': item['idrecurso__nombrerecurso'],
            'total': item['total'],
            'disponible': item['disponible'],
            'reparacion': item['reparacion'],
            'fuera_servicio': item['fuera_servicio']
        }
        for item in recursos_data
    }

    context = {
        'recursos': recursos_dict,
        'total_recursos': len(recursos_dict)
    }
    
    return render(request, 'recursos.html', context)

def login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)

        if user is not None:
            auth_login(request, user)
            request.session["access_token"] = "aquí_va_tu_token"
            return redirect("/panel")
        else:
            return render(request, "login.html", {"error": "Usuario o contraseña incorrectos"})
    return render(request, "login.html")

@login_required(login_url='/login')
def boxes(request):
    return render(request, 'boxes.html')

@login_required(login_url='/login')
def panel_admin(request):
    return render(request, 'panel_admin.html')

@login_required(login_url='/login')
def dashboard(request):
    hoy = date.today()
    ahora = datetime.now().time()
    
    total_espacios = Espacio.objects.count()
    
    en_mantencion = Espacio.objects.filter(idestadoespacio__descripcionestadoespacio="En mantención").count()
    
    espacios_en_uso_ids = Reserva.objects.filter(
        fechareserva=hoy,
        horainicio__lte=ahora,
        horafin__gte=ahora
    ).values_list('idespacio', flat=True).distinct()
    
    espacios_en_uso = espacios_en_uso_ids.count()
    
    espacios_en_uso_y_mantencion = Espacio.objects.filter(
        idespacio__in=espacios_en_uso_ids,
        idestadoespacio__descripcionestadoespacio="En mantención"
    ).count()
    
    espacios_realmente_disponibles = total_espacios - espacios_en_uso - en_mantencion + espacios_en_uso_y_mantencion
    
    disponibles = max(0, espacios_realmente_disponibles)
    
    en_mantencion = max(0, en_mantencion - espacios_en_uso_y_mantencion)
    
    porcentajes = {
        'en_uso': round((espacios_en_uso / total_espacios * 100), 1) if total_espacios > 0 else 0,
        'disponibles': round((disponibles / total_espacios * 100), 1) if total_espacios > 0 else 0,
        'en_mantencion': round((en_mantencion / total_espacios * 100), 1) if total_espacios > 0 else 0,
    }

    horas = ['08:00','08:20','08:40','09:00','09:20','09:40','10:00','10:20','10:40', 
             '11:00', '11:20', '11:40', '12:00','12:20','12:40','13:00', '13:20', '13:40', '14:00','14:20', '14:40', 
             '15:00','15:20','15:40','16:00', '16:20', '16:40', '17:00','17:20','17:40']
    ocupados_por_hora = []
    
    for hora_str in horas:
        hora = datetime.strptime(hora_str, '%H:%M').time()
        espacios_en_uso_hora = Reserva.objects.filter(
            fechareserva=hoy,
            horainicio__lte=hora,
            horafin__gte=hora
        ).values_list('idespacio', flat=True).distinct().count()
        ocupados_por_hora.append(espacios_en_uso_hora)

    tipos_actividad_demandadas = Reserva.objects.filter(fechareserva=hoy).values(
        'rutresponsable__idtipoactividad__nombretipoactividad'
    ).annotate(
        total=Count('idreserva')
    ).order_by('-total')[:5]

    labels_tipos_actividad = [e['rutresponsable__idtipoactividad__nombretipoactividad'] for e in tipos_actividad_demandadas]
    data_tipos_actividad = [e['total'] for e in tipos_actividad_demandadas]

    context = {
        'total_espacios': total_espacios,
        'en_uso': espacios_en_uso,
        'disponibles': disponibles,
        'en_mantencion': en_mantencion,
        'porcentajes': porcentajes,
        'horas': json.dumps(horas),
        'ocupados_por_hora': json.dumps(ocupados_por_hora),
        'labels_tipos_actividad': json.dumps(labels_tipos_actividad),
        'data_tipos_actividad': json.dumps(data_tipos_actividad),
    }
    return render(request, 'dashboard.html', context)

@login_required(login_url='/login')
def box(request):
    return render(request, 'box.html')

def exportar_excel_box(request, idbox):
    fecha_str = request.GET.get('fecha')
    try:
        fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        fecha = date.today()

    consultas = Consulta.objects.filter(
        idbox=idbox,
        fechaconsulta=fecha
    ).select_related('rutmedico', 'rutpaciente', 'idbox', 'idtipocita')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Turnos Box"
    
    ws.append(['Box', 'Médico', 'Paciente', 'Fecha', 'Hora', 'Tipo de Cita'])
    
    for consulta in consultas:
        medico = ""
        if consulta.rutmedico:
            nombre_medico = getattr(consulta.rutmedico, "nombremedico", "") or getattr(consulta.rutmedico, "nombre", "")
            apellido_medico = getattr(consulta.rutmedico, "apellidomedico", "") or getattr(consulta.rutmedico, "apellido", "")
            medico = f"{nombre_medico} {apellido_medico}".strip()

        paciente = ""
        if consulta.rutpaciente:
            nombre_paciente = getattr(consulta.rutpaciente, "nombrepaciente", "") or getattr(consulta.rutpaciente, "nombre", "")
            apellido_paciente = getattr(consulta.rutpaciente, "apellidopaciente", "") or getattr(consulta.rutpaciente, "apellido", "")
            paciente = f"{nombre_paciente} {apellido_paciente}".strip()

        tipo_cita = getattr(consulta.idtipocita, "tipocita", "")

        ws.append([
            f"Box {consulta.idbox.numerobox if hasattr(consulta.idbox, 'numerobox') else consulta.idbox_id}",
            medico,
            paciente,
            consulta.fechaconsulta.strftime('%Y-%m-%d'),
            consulta.horainicio.strftime('%H:%M'),
            tipo_cita,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"box_{idbox}_turnos_{fecha}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def exportar_excel_espacio(request, idespacio):
    fecha_str = request.GET.get('fecha')
    try:
        fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        fecha = date.today()

    reservas = Reserva.objects.filter(
        idespacio=idespacio,
        fechareserva=fecha
    ).select_related('rutresponsable', 'rutusuario', 'idespacio', 'idtiporeserva')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reservas Espacio"
    
    ws.append(['Espacio', 'Responsable', 'Usuario', 'Fecha', 'Hora', 'Tipo de Reserva'])
    
    for reserva in reservas:
        responsable = ""
        if reserva.rutresponsable:
            nombre_responsable = getattr(reserva.rutresponsable, "nombreresponsable", "") or getattr(reserva.rutresponsable, "nombre", "")
            apellido_responsable = getattr(reserva.rutresponsable, "apellidoresponsable", "") or getattr(reserva.rutresponsable, "apellido", "")
            responsable = f"{nombre_responsable} {apellido_responsable}".strip()

        usuario = ""
        if reserva.rutusuario:
            nombre_usuario = getattr(reserva.rutusuario, "nombreusuario", "") or getattr(reserva.rutusuario, "nombre", "")
            apellido_usuario = getattr(reserva.rutusuario, "apellidousuario", "") or getattr(reserva.rutusuario, "apellido", "")
            usuario = f"{nombre_usuario} {apellido_usuario}".strip()

        tipo_reserva = getattr(reserva.idtiporeserva, "tiporeserva", "")

        ws.append([
            f"Espacio {reserva.idespacio.numeroespacio if hasattr(reserva.idespacio, 'numeroespacio') else reserva.idespacio_id}",
            responsable,
            usuario,
            reserva.fechareserva.strftime('%Y-%m-%d'),
            reserva.horainicio.strftime('%H:%M'),
            tipo_reserva,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"espacio_{idespacio}_reservas_{fecha}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def homepage(request):
    return render(request, 'homepage.html')

def agendar_general(request):
    if request.method == "POST":
        try:
            fecha = request.POST.get("fechaReserva")
            hora_inicio = request.POST.get("horaInicio")
            hora_fin = request.POST.get("horaFin")
            espacio_id = request.POST.get("espacio")
            tipo_reserva_id = 1  # Tipo general

            if not (fecha and hora_inicio and hora_fin and espacio_id):
                return JsonResponse({"mensaje": "Faltan datos para agendar"}, status=400)

            choque = Reserva.objects.filter(
                idespacio=espacio_id,
                fechareserva=fecha
            ).filter(
                Q(horainicio__lt=hora_fin) & Q(horafin__gt=hora_inicio)
            ).exists()

            if choque:
                return JsonResponse({"mensaje": "El espacio tiene una reserva agendada en este horario"}, status=400)

            espacio = Espacio.objects.get(pk=espacio_id)
            tipo_reserva = TipoReserva.objects.get(pk=tipo_reserva_id)

            Reserva.objects.create(
                idespacio=espacio,
                fechareserva=fecha,
                horainicio=hora_inicio,
                horafin=hora_fin,
                idtiporeserva=tipo_reserva
            )

            return JsonResponse({"mensaje": "La reserva ha sido agendada correctamente"})

        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=400)

def agendar_no_medica(request):
    if request.method == "POST":
        try:
            fecha = request.POST.get("fechaConsulta")
            hora_inicio = request.POST.get("horaInicio")
            hora_fin = request.POST.get("horaFin")
            box_id = request.POST.get("box")
            tipo_cita_id = 2

            if not (fecha and hora_inicio and hora_fin and box_id):
                return JsonResponse({"mensaje": "Faltan datos para agendar"}, status=400)

            choque = Consulta.objects.filter(
                idbox=box_id,
                fechaconsulta=fecha
            ).filter(
                Q(horainicio__lt=hora_fin) & Q(horafin__gt=hora_inicio)
            ).exists()

            if choque:
                return JsonResponse({"mensaje": "El box tiene una cita agendada en este horario"}, status=400)

            box = Box.objects.get(pk=box_id)
            tipo_cita = TipoCita.objects.get(pk=tipo_cita_id)

            Consulta.objects.create(
                idbox=box,
                fechaconsulta=fecha,
                horainicio=hora_inicio,
                horafin=hora_fin,
                idtipocita=tipo_cita
            )

            return JsonResponse({"mensaje": "La cita ha sido agendada correctamente"})

        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=400)
            
def agendar_general_espacio(request, espacio_id):
    if request.method == "POST":
        try:
            fecha = request.POST.get("fechaReserva")
            hora_inicio = request.POST.get("horaInicio")
            hora_fin = request.POST.get("horaFin")
            tipo_reserva_id = 1  # Tipo general

            if not (fecha and hora_inicio and hora_fin):
                return JsonResponse({"mensaje": "Faltan datos para agendar"}, status=400)

            choque = Reserva.objects.filter(
                idespacio=espacio_id,
                fechareserva=fecha
            ).filter(
                Q(horainicio__lt=hora_fin) & Q(horafin__gt=hora_inicio)
            ).exists()

            if choque:
                return JsonResponse({"mensaje": "El espacio tiene una reserva agendada en este horario"}, status=400)

            espacio = Espacio.objects.get(pk=espacio_id)
            tipo_reserva = TipoReserva.objects.get(pk=tipo_reserva_id)

            Reserva.objects.create(
                idespacio=espacio,
                fechareserva=fecha,
                horainicio=hora_inicio,
                horafin=hora_fin,
                idtiporeserva=tipo_reserva
            )

            return JsonResponse({"mensaje": "La reserva ha sido agendada correctamente"})
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=400)

def agendar_no_medica_box(request, box_id):
    if request.method == "POST":
        try:
            fecha = request.POST.get("fechaConsulta")
            hora_inicio = request.POST.get("horaInicio")
            hora_fin = request.POST.get("horaFin")
            tipo_cita_id = 2

            if not (fecha and hora_inicio and hora_fin):
                return JsonResponse({"mensaje": "Faltan datos para agendar"}, status=400)

            choque = Consulta.objects.filter(
                idbox=box_id,
                fechaconsulta=fecha
            ).filter(
                Q(horainicio__lt=hora_fin) & Q(horafin__gt=hora_inicio)
            ).exists()

            if (choque):
                return JsonResponse({"mensaje": "El box tiene una cita agendada en este horario"}, status=400)

            box = Box.objects.get(pk=box_id)
            tipo_cita = TipoCita.objects.get(pk=tipo_cita_id)

            Consulta.objects.create(
                idbox=box,
                fechaconsulta=fecha,
                horainicio=hora_inicio,
                horafin=hora_fin,
                idtipocita=tipo_cita
            )

            return JsonResponse({"mensaje": "La cita ha sido agendada correctamente"})
        except Exception as e:
            return JsonResponse({"mensaje": f"Error: {str(e)}"}, status=400)


def reporte_por_dia(request):
    if request.method == 'POST':
        accion = request.POST.get('accion')

        if accion == 'descargar':
            fecha_str = request.POST.get('fecha')
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()

            consultas = Consulta.objects.filter(fechaconsulta=fecha)

            wb = Workbook()
            ws = wb.active
            ws.title = f"Consultas {fecha_str}"

            ws.append(['ID', 'Paciente', 'Médico', 'Box', 'Hora Inicio', 'Hora Fin', 'Tipo de cita'])

            for consulta in consultas:
                paciente = (
                    f"{consulta.rutpaciente.nombrepaciente} {consulta.rutpaciente.apellidopaciente}"
                    if consulta.rutpaciente else "Sin paciente"
                )
                medico = (
                    f"{consulta.rutmedico.nombremedico} {consulta.rutmedico.apellidomedico}"
                    if consulta.rutmedico else "Sin médico"
                )
                box = consulta.idbox.numerobox if consulta.idbox else "Sin box"
                hora_inicio = str(consulta.horainicio) if consulta.horainicio else ""
                hora_fin = str(consulta.horafin) if consulta.horafin else ""
                tipo_cita = consulta.idtipocita.tipocita if consulta.idtipocita else "Sin tipo"

                ws.append([
                    consulta.idconsulta,
                    paciente,
                    medico,
                    box,
                    hora_inicio,
                    hora_fin,
                    tipo_cita
                ])

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'reporte_{fecha_str}.xlsx'
            response['Content-Disposition'] = f'attachment; filename={filename}'
            wb.save(response)

            return response

        elif accion == 'importar' and request.FILES.get('archivo'):
            archivo = request.FILES['archivo']

            try:
                if archivo.name.endswith('.csv'):
                    df = pd.read_csv(archivo)
                else:
                    df = pd.read_excel(archivo)
            except Exception as e:
                messages.error(request, f"Error al leer el archivo: {e}")
                return redirect('reporte_por_dia')

            errores = []
            for i, row in df.iterrows():
                fecha = row['fecha']
                hora_inicio = row['horaInicio']
                hora_fin = row['horaFin']
                box_id = row['idBox']

                choques = Consulta.objects.filter(
                    idbox_id=box_id,
                    fechaconsulta=fecha,
                    horainicio__lt=hora_fin,
                    horafin__gt=hora_inicio
                )

                if choques.exists():
                    errores.append(f"Fila {i+2}: Conflicto en Box {box_id} entre {hora_inicio}-{hora_fin}")
                else:
                    Consulta.objects.create(
                        idbox_id=box_id,
                        rutpaciente_id=row['rutPaciente'],
                        rutmedico_id=row['rutMedico'],
                        fechaconsulta=fecha,
                        horainicio=hora_inicio,
                        horafin=hora_fin,
                        idtipocita_id=row.get('idTipoCita')
                    )

            if errores:
                for err in errores:
                    messages.error(request, err)
            else:
                messages.success(request, "Horario importado correctamente.")

            return redirect('reportes')

    return render(request, 'reportes.html')

def ajax_reporte_preview(request):
    fecha_str = request.GET.get('fecha')
    try:
        fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return JsonResponse({'columns': [], 'rows': []})

    consultas = Consulta.objects.filter(fechaconsulta=fecha)

    columns = ['ID', 'Paciente', 'Médico', 'Box', 'Hora Inicio', 'Hora Fin', 'Tipo de cita']
    rows = []

    for consulta in consultas:
        paciente = f"{consulta.rutpaciente.nombrepaciente} {consulta.rutpaciente.apellidopaciente}" if consulta.rutpaciente else "Sin paciente"
        medico = f"{consulta.rutmedico.nombremedico} {consulta.rutmedico.apellidomedico}" if consulta.rutmedico else "Sin médico"
        box = consulta.idbox.numerobox if consulta.idbox else "Sin box"
        hora_inicio = str(consulta.horainicio) if consulta.horainicio else ""
        hora_fin = str(consulta.horafin) if consulta.horafin else ""
        tipo_cita = consulta.idtipocita.tipocita if consulta.idtipocita else "Sin tipo"

        rows.append([consulta.idconsulta, paciente, medico, box, hora_inicio, hora_fin, tipo_cita])

    return JsonResponse({'columns': columns, 'rows': rows})


def ajax_importar_preview(request):
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        try:
            if archivo.name.endswith('.csv'):
                df = pd.read_csv(archivo)
            else:
                df = pd.read_excel(archivo)
        except:
            return JsonResponse({'columns': [], 'rows': []})

        columns = list(df.columns)
        rows = df.head(10).values.tolist()
        return JsonResponse({'columns': columns, 'rows': rows})
    return JsonResponse({'columns': [], 'rows': []})